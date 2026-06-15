import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# 1. Configuración de URLs y Credenciales de la API local
API_URL = "http://localhost:5267/api"
AUTH_PAYLOAD = {
    "usuario": "admin",
    "password": "admin123",
    "tenantId": 1  # First National Bank
}

print("🔐 Autenticando en NeboAudit API...")
try:
    # 2. Obtención dinámica del Token JWT
    auth_response = requests.post(f"{API_URL}/auth/login", json=AUTH_PAYLOAD)
    auth_response.raise_for_status()
    token = auth_response.json().get("token")
    
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json"
    }

    # 3. Consumo del endpoint de Inteligencia y Riesgo Crítico
    print("📊 Extrayendo portafolio de alto riesgo (RiskScore >= 0.5000)...")
    loans_response = requests.get(f"{API_URL}/Loans/high-risk?threshold=0.5000", headers=headers)
    loans_response.raise_for_status()
    loans_data = loans_response.json()

    if not loans_data:
        print("⚠️ No se encontraron créditos de alto riesgo para el Tenant seleccionado.")
        exit()

    # 4. Procesamiento de Big Data con Pandas
    df = pd.DataFrame(loans_data)
    # Seleccionar y reordenar columnas clave para el Comité de Riesgos
    columns_order = [
        'externalLoanId', 'loanType', 'loanAmount', 'propertyValue', 
        'ltv', 'dti', 'riskScore', 'automatedAction', 'auditDays'
    ]
    df = df[columns_order]

    # 5. Diseño Estético del Reporte Excel Corporativo (openpyxl)
    wb = Workbook()
    ws = wb.active
    ws.title = "Comité de Riesgo Hipotecario"
    ws.views.sheetView[0].showGridLines = True

    # Estilos de diseño
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Azul Marino
    
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Alerta Crítica
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Revisión Senior
    
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # Inyectar DataFrame al Excel
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Aplicar formatos y reglas de negocio visuales
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Formatear cabeceras
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Formatear celdas operativas y aplicar Formato Condicional
    for row in range(2, ws.max_row + 1):
        # Alineaciones y formatos numéricos financieros
        ws[f'A{row}'].alignment = center_align
        ws[f'B{row}'].alignment = center_align
        ws[f'C{row}'].number_format = '$#,##0.00'
        ws[f'D{row}'].number_format = '$#,##0.00'
        ws[f'E{row}'].number_format = '0.00"%"'
        ws[f'F{row}'].number_format = '0.00"%"'
        ws[f'G{row}'].number_format = '0.0000'
        ws[f'H{row}'].alignment = center_align
        ws[f'I{row}'].number_format = '0.0'

        # Evaluar la acción automatizada dictada por la IA corporativa
        action = ws[f'H{row}'].value
        if action == "AUTO_BLOCK":
            ws[f'H{row}'].fill = red_fill
            ws[f'G{row}'].fill = red_fill # Pintar también el score de riesgo
        elif action == "SENIOR_REVIEW":
            ws[f'H{row}'].fill = yellow_fill
            ws[f'G{row}'].fill = yellow_fill

    # Guardar reporte finalizado
    output_filename = "NeboAudit_Risk_Report.xlsx"
    wb.save(output_filename)
    print(f"🎉 ¡Reporte analítico generado exitosamente! Archivo guardado como: {output_filename}")

except Exception as e:
    print(f"❌ Error crítico en el pipeline de datos: {e}")